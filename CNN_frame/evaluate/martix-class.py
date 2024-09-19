# -*- coding: utf-8 -*-
import os
import re
import numpy as np
import openpyxl
from PIL import Image
import numpy as np
from sklearn.metrics import confusion_matrix
import pandas as pd

# 计算OA、Kappa等指标的函数
def calculate_oa(cm):
    total_correct = np.trace(cm)
    total_samples = np.sum(cm)
    oa = total_correct / total_samples
    return oa

def calculate_kappa(oa, cm):
    total_samples = np.sum(cm)
    pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total_samples ** 2)
    kappa = (oa - pe) / (1 - pe)
    return kappa

def calculate_binary_confusion_matrices(matrix):
    num_classes = matrix.shape[0]
    binary_confusion_matrices = []

    for class_idx in range(num_classes):
        true_positive = matrix[class_idx, class_idx]
        false_positive = np.sum(matrix[:, class_idx]) - true_positive
        false_negative = np.sum(matrix[class_idx, :]) - true_positive
        true_negative = np.sum(matrix) - (true_positive + false_positive + false_negative)

        binary_matrix = np.array([[true_positive, false_positive],
                                  [false_negative, true_negative]])

        binary_confusion_matrices.append(binary_matrix)

    return binary_confusion_matrices

def calculate_precision_recall_f1(cm):
    num_classes = len(cm)
    precision = []
    recall = []
    f1_score = []

    for i in range(num_classes):
        tp = cm[i, i]
        fp = np.sum(cm[:, i]) - tp
        fn = np.sum(cm[i, :]) - tp

        precision.append(tp / (tp + fp))
        recall.append(tp / (tp + fn))
        f1_score.append(2 * (precision[i] * recall[i]) / (precision[i] + recall[i]))

    return precision, recall, f1_score


# 加载PNG图像并转换为NumPy数组
predicted_image = np.array(Image.open(r"E:\Train\Train\Rural\mask_low\0.png"))
label_image = np.array(Image.open(r"E:\Train\Train\Rural\mask_low\0.png"))

# 检查图像尺寸是否一致
if predicted_image.shape != label_image.shape:
    raise ValueError("预测图和标签图的尺寸不一致")

# 将预测图和标签图展平为一维数组
predicted_labels = predicted_image.flatten()
true_labels = label_image.flatten()

# 计算混淆矩阵
cm = confusion_matrix(true_labels, predicted_labels)

# 保存混淆矩阵
np.save('path_to_save_confusion_matrix.npy', cm)

print("混淆矩阵已生成并保存。")


folder_path = r"E:\Train\Train\1"

# 遍历文件夹中的所有文件
for file_name in os.listdir(folder_path):
    if file_name.endswith(".npy"):
        print(file_name)
        # 读取Numpy数组
        matrix = np.load(os.path.join(folder_path, file_name))

        # 计算混淆矩阵等指标
        binary_matrices = calculate_binary_confusion_matrices(matrix)
        num_classes = matrix.shape[0]
        producer_accuracy = [matrix[i, i] / np.sum(matrix[:, i]) for i in range(num_classes)]
        user_accuracy = [matrix[i, i] / np.sum(matrix[i, :]) for i in range(num_classes)]
        overall_accuracy = calculate_oa(matrix)
        kappa = calculate_kappa(overall_accuracy, matrix)

        # 输出指标
        print(f"整体kappa：{kappa:.3f} and 整体oa：{overall_accuracy:.3f} ")

        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 写入行标题
        sheet['A1'] = "Metrics"
        metrics = ["Kappa", "OA", "UA", "PA", "F1 Score"]
        for row_idx, metric in enumerate(metrics, start=2):
            sheet.cell(row=row_idx, column=1, value=metric)

        # 写入列标题
        sheet['B1'] = "Class"
        for col_idx in range(2, num_classes + 2):
            sheet.cell(row=1, column=col_idx, value=f'Class {col_idx - 2}')

        # 遍历每个类别，写入数据
        for class_idx in range(num_classes):
            sheet.cell(row=1, column=class_idx + 2, value=f'Class {class_idx}')
            precision, recall, f1_score = calculate_precision_recall_f1(matrix)

            for row_idx, metric in enumerate(metrics, start=2):
                if metric == "Kappa":
                    value = calculate_kappa(calculate_oa(binary_matrices[class_idx]), binary_matrices[class_idx])
                elif metric == "OA":
                    value = calculate_oa(binary_matrices[class_idx])
                elif metric == "UA":
                    value = user_accuracy[class_idx]
                elif metric == "PA":
                    value = producer_accuracy[class_idx]
                elif metric == "F1 Score":
                    value = f1_score[class_idx]
                else:
                    value = None  # 在这里添加其他指标的处理
                sheet.cell(row=row_idx, column=class_idx + 2, value=value)
        # 总体精度
        overall_accuracy = calculate_oa(matrix)
        sheet['L1'] = "总体精度"
        sheet['M1'] = overall_accuracy

        # 卡帕系数
        kappa = calculate_kappa(overall_accuracy, matrix)
        sheet['L2'] = "卡帕系数"
        sheet['M2'] = kappa
        # 保存工作簿为Excel文件
        output_name = file_name
        excel_file_path = os.path.join(folder_path, f'{output_name}.xlsx')
        workbook.save(excel_file_path)

        print(f"Excel文件已创建并保存：'{excel_file_path}'")


